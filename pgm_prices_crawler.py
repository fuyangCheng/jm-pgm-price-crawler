# -*- coding: utf-8 -*-
"""
Johnson Matthey PGM 价格抓取脚本（修正版）
- 所有数据与网站保持一致：日期、时间、价格等均来自网页 JSON，不做本地覆盖或默认值。
- 统一使用 Playwright 解析隐藏 JSON（currentMetalPrices / metalTablePrices）
- Platinum：写入 4 个时段（Fix），日期与分时价格均取自 metalTablePrices
- Palladium / Rhodium：当前价与日期取自 currentMetalPrices，时段列与网站展示一致
- Date：使用 JM 返回的有效日期/更新时间（与网站显示一致），仅当接口未返回日期时用抓取日
- Excel：按金属分 Sheet，列为 Date, Source, Hong Kong 08:30, Hong Kong 14:00, London 09:00, New York 09:30
  合并键：Date（每个金属每天一行）
- 单次模式（--once）与交互模式使用同一抓取链路

运行方式
--------
  python pgm_prices_crawler.py           # 交互模式（定时抓取 + 输入指令）
  python pgm_prices_crawler.py --once    # 单次抓取，不进入交互，执行完即退出

交互模式下指令
-------------
  status  查看运行状态（香港时间、最近抓取记录、缺失项、下次抓取倒计时）
  view    打开 all_metals_data.xlsx
  force   一键抓取全部地区/时段（Pt 四时段 + Pd/Rh 当前价填四列），无需选择
  exit    安全退出
  help    显示帮助
"""

import os
import re
import sys
import json
import time
import threading
import subprocess
from datetime import datetime, timedelta
from typing import Optional, Tuple
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from openpyxl.chart import LineChart, Reference

# Playwright 用于 JS 渲染/读取隐藏 input JSON；未安装则无法使用主链路
try:
    from playwright.sync_api import sync_playwright
    _PLAYWRIGHT_AVAILABLE = True
except ImportError:
    _PLAYWRIGHT_AVAILABLE = False

# ---------- 配置 ----------
DATA_SOURCE_URL = "https://matthey.com/products-and-markets/pgms-and-circularity/pgm-management"
OUTPUT_FILE = "all_metals_data.xlsx"

PLAYWRIGHT_PAGE_TIMEOUT_MS = 60000
PLAYWRIGHT_SELECTOR_TIMEOUT_MS = 25000

DEBUG_SAVE_RAW_TABLE = True  # 保存调试 JSON：PGM_debug_currentMetalPrices.json / PGM_debug_metalTablePrices.json / PGM_debug_parsed_meta.json
# 与网站保持一致：为 False 时 Date 使用 JM 返回的有效日期；为 True 时用抓取当天（不推荐，会与网站显示不一致）
USE_TODAY_AS_DATE = False

TARGET_METALS = ("Platinum", "Palladium", "Rhodium")
TIME_COLUMNS = [
    "Hong Kong 08:30",
    "Hong Kong 14:00",
    "London 09:00",
    "New York 09:30",
]

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# ---------- 多线程与调度 ----------
_state_lock = threading.Lock()
_latest_scraped_record = None  # dict，用于 status
_next_run_time = None  # datetime (HKT)
_latest_missing_items = []  # 最近一次抓取中缺失的具体金属/时段
_shutdown_event = threading.Event()
_excel_lock = threading.Lock()

WAIT_RETRY_MINUTES = 5
WAIT_DEADLINE_HOURS = 1


def _log(msg: str) -> None:
    ts = datetime.now(ZoneInfo("Asia/Hong_Kong")).strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}")


def _to_float(val) -> Optional[float]:
    """将价格转为浮点数，支持数字或带 $/逗号的字符串。"""
    if val is None or val == "" or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)):
        try:
            f = float(val)
            return f if f == f else None  # reject nan
        except (TypeError, ValueError):
            return None
    s = str(val).strip().replace(",", "").replace("$", "").replace(" ", "")
    if s in ("-", "--", "—", "nan", ""):
        return None
    s = re.sub(r"\s*/\s*.*$", "", s)
    try:
        return float(s)
    except ValueError:
        return None


def _current_session_column() -> Optional[str]:
    """
    根据真实时区换算当前香港时间所属的交易时段（自动处理夏令时）：
    - Hong Kong 08:30 / 14:00 按香港本地时间判断
    - London 09:00 / New York 09:30 通过 Europe/London、America/New_York 转换
    """
    hkt = ZoneInfo("Asia/Hong_Kong")
    now_hkt = datetime.now(hkt)

    london_tz = ZoneInfo("Europe/London")
    ny_tz = ZoneInfo("America/New_York")

    today_hkt = now_hkt.date()
    hk0830 = datetime(today_hkt.year, today_hkt.month, today_hkt.day, 8, 30, tzinfo=hkt)
    hk1400 = datetime(today_hkt.year, today_hkt.month, today_hkt.day, 14, 0, tzinfo=hkt)

    now_london = datetime.now(london_tz)
    london_0900_local = now_london.replace(hour=9, minute=0, second=0, microsecond=0)
    london_0900_hkt = london_0900_local.astimezone(hkt)

    now_ny = datetime.now(ny_tz)
    ny_0930_local = now_ny.replace(hour=9, minute=30, second=0, microsecond=0)
    ny_0930_hkt = ny_0930_local.astimezone(hkt)

    window_min = 60

    def in_window(t: datetime) -> bool:
        return t <= now_hkt < (t + timedelta(minutes=window_min))

    if in_window(hk0830):
        return "Hong Kong 08:30"
    if in_window(hk1400):
        return "Hong Kong 14:00"
    if in_window(london_0900_hkt):
        return "London 09:00"
    if in_window(ny_0930_hkt):
        return "New York 09:30"
    return None


def _get_next_run_time_hkt() -> datetime:
    """
    返回下一次应执行抓取的香港时间（HKT），使用真实时区换算，自动处理伦敦/纽约夏令时。
    """
    hkt = ZoneInfo("Asia/Hong_Kong")
    now_hkt = datetime.now(hkt)

    london_tz = ZoneInfo("Europe/London")
    ny_tz = ZoneInfo("America/New_York")

    candidates = []

    # 香港本地两次固定
    for hh, mm in [(8, 30), (14, 0)]:
        t = now_hkt.replace(hour=hh, minute=mm, second=0, microsecond=0)
        if t <= now_hkt:
            t = t + timedelta(days=1)
        candidates.append(t)

    # London 09:00 -> HKT
    now_london = datetime.now(london_tz)
    london_0900_local = now_london.replace(hour=9, minute=0, second=0, microsecond=0)
    london_0900_hkt = london_0900_local.astimezone(hkt)
    if london_0900_hkt <= now_hkt:
        london_0900_hkt = (london_0900_local + timedelta(days=1)).astimezone(hkt)
    candidates.append(london_0900_hkt)

    # NY 09:30 -> HKT
    now_ny = datetime.now(ny_tz)
    ny_0930_local = now_ny.replace(hour=9, minute=30, second=0, microsecond=0)
    ny_0930_hkt = ny_0930_local.astimezone(hkt)
    if ny_0930_hkt <= now_hkt:
        ny_0930_hkt = (ny_0930_local + timedelta(days=1)).astimezone(hkt)
    candidates.append(ny_0930_hkt)

    return min(candidates)


def _session_name_for_run_time(next_t_hkt: datetime) -> Optional[str]:
    """根据下一次触发时间（HKT）反推所属的 session 名称，用于写入正确的时段列。"""
    london_tz = ZoneInfo("Europe/London")
    ny_tz = ZoneInfo("America/New_York")

    h, m = next_t_hkt.hour, next_t_hkt.minute
    if (h, m) == (8, 30):
        return "Hong Kong 08:30"
    if (h, m) == (14, 0):
        return "Hong Kong 14:00"

    london_time = next_t_hkt.astimezone(london_tz)
    if london_time.hour == 9 and london_time.minute == 0:
        return "London 09:00"

    ny_time = next_t_hkt.astimezone(ny_tz)
    if ny_time.hour == 9 and ny_time.minute == 30:
        return "New York 09:30"

    return None


def _extract_effective_date(obj) -> Optional[str]:
    """
    从 JSON 中提取与网站显示一致的日期（YYYY-MM-DD）。
    优先使用「价格日期」类字段；若为带时区的时间戳，按欧洲/伦敦取日期以与 JM 网站一致。
    """
    if not isinstance(obj, dict):
        return None
    # 优先价格/生效日期，避免用「更新时间」导致比网站多一天
    candidate_keys = [
        "priceDate",
        "effectiveDate",
        "asOfDate",
        "asOf",
        "dateStr",
        "date",
        "asOfTime",
        "timestamp",
        "updated",
        "updateTime",
    ]
    london_tz = ZoneInfo("Europe/London")
    for k in candidate_keys:
        if k not in obj or obj[k] in (None, ""):
            continue
        s = str(obj[k]).strip()
        if not s:
            continue
        try:
            dt = pd.to_datetime(s, errors="coerce", utc=True)
            if pd.isna(dt):
                dt = pd.to_datetime(s[:10], errors="coerce", utc=True) if len(s) >= 10 else pd.NaT
            if pd.isna(dt):
                continue
            # 与网站一致：JM 显示为英国日期，将时间戳转为伦敦日期
            if hasattr(dt, "tz") and dt.tz is not None:
                dt_ldn = dt.tz_convert(london_tz)
            else:
                dt_ldn = pd.Timestamp(dt).tz_localize("UTC").tz_convert(london_tz)
            return dt_ldn.strftime("%Y-%m-%d")
        except Exception:
            try:
                # 纯日期字符串如 2025-02-27，直接取前 10 位
                if len(s) >= 10:
                    dt = pd.to_datetime(s[:10], errors="coerce")
                    if not pd.isna(dt):
                        return dt.strftime("%Y-%m-%d")
            except Exception:
                continue
    return None


def _scrape_rows_with_playwright(session_override: Optional[str] = None) -> pd.DataFrame:
    """
    使用 Playwright 解析 JM 页面隐藏 input(JSON)：
    - currentMetalPrices/allCurrentMetalPrices: Pt/Pd/Rh 当前价
    - metalTablePrices: Pt 分市场/分时段价格（通常是 Fix）
    输出列: Date, Metal, Source, 4 个 TIME_COLUMNS
    """
    if not _PLAYWRIGHT_AVAILABLE:
        raise RuntimeError("未安装 Playwright。请执行: pip install playwright  然后: playwright install chromium")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            page = browser.new_page()
            page.set_default_timeout(PLAYWRIGHT_SELECTOR_TIMEOUT_MS)
            page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
            page.goto(DATA_SOURCE_URL, wait_until="commit", timeout=PLAYWRIGHT_PAGE_TIMEOUT_MS)

            try:
                page.wait_for_selector(
                    "#currentMetalPrices, #allCurrentMetalPrices, table#priceTableHTML, "
                    "#_jm_metal_price_table_portlet_MetalPriceTablePortlet_metalTablePrices",
                    state="attached",
                    timeout=PLAYWRIGHT_SELECTOR_TIMEOUT_MS,
                )
            except Exception:
                page.wait_for_selector("table, input", state="attached", timeout=10000)
            page.wait_for_timeout(2000)

            today_hkt = datetime.now(ZoneInfo("Asia/Hong_Kong")).strftime("%Y-%m-%d")
            _log(f"本次抓取 session_override={session_override}, 推断写入列={_current_session_column()}")

            # 1) currentMetalPrices
            cur_json = page.locator("#currentMetalPrices").get_attribute("value")
            if not cur_json:
                cur_json = (
                    page.locator("#allCurrentMetalPrices").get_attribute("value")
                    or page.locator("#_jm_metal_price_table_portlet_MetalPriceTablePortlet_allCurrentMetalPrices").get_attribute("value")
                )

            cur_data = json.loads(cur_json) if cur_json else {}
            cur_list = cur_data.get("currentMetalList", cur_data.get("metalList", []))
            cur_effective_dt = _extract_effective_date(cur_data)

            # 2) metalTablePrices (Pt)
            table_json = page.locator("#_jm_metal_price_table_portlet_MetalPriceTablePortlet_metalTablePrices").get_attribute("value")
            pt_time_prices = {tc: None for tc in TIME_COLUMNS}
            pt_effective_dt: Optional[str] = None

            def _norm_market(s: str) -> str:
                if not s:
                    return ""
                s = str(s).strip().lower()
                if s in ("hong kong", "hk"):
                    return "Hong Kong"
                if s in ("london", "ldn"):
                    return "London"
                if s in ("new york", "ny", "newyork"):
                    return "New York"
                return str(s).strip()

            def _norm_time(s: str) -> str:
                if s is None:
                    return ""
                s = str(s).strip().replace(" ", "")
                if "08:30" in s or "8:30" in s:
                    return "08:30"
                if "14:00" in s or "2:00" in s:
                    return "14:00"
                if "09:00" in s or "9:00" in s:
                    return "09:00"
                if "09:30" in s or "9:30" in s:
                    return "09:30"
                return s

            def _items_from_group(g) -> Tuple[Optional[str], list]:
                """从一组数据中取出日期和 4 个市场项。支持 list 或 dict（dict 时取 .prices / .markets / .priceList）。"""
                group_date: Optional[str] = None
                items: list = []
                if isinstance(g, list):
                    items = [x for x in g if isinstance(x, dict)]
                    for it in items:
                        if group_date is None:
                            group_date = _extract_effective_date(it)
                elif isinstance(g, dict):
                    group_date = _extract_effective_date(g)
                    items = g.get("prices") or g.get("markets") or g.get("priceList") or g.get("data") or []
                    if not isinstance(items, list):
                        items = []
                    items = [x for x in items if isinstance(x, dict)]
                return group_date, items

            def pick_latest_valid_group(groups: list) -> Optional[list]:
                """从 groups 中选“日期最新 + 有效值最多”的组，返回该组的 4 项 list。"""
                best_idx = None
                best_dt = None
                best_valid = -1
                flattened: list = []  # 最终返回的 list of dicts

                def parse_dt(s: Optional[str]) -> Optional[datetime]:
                    if not s:
                        return None
                    try:
                        dt = pd.to_datetime(s, errors="coerce")
                        if pd.isna(dt):
                            return None
                        return dt.to_pydatetime()
                    except Exception:
                        return None

                for idx, g in enumerate(groups):
                    group_date, items = _items_from_group(g)
                    valid_count = sum(
                        1 for it in items
                        if _to_float(it.get("price") or it.get("value") or it.get("currentPrice")) is not None
                    )
                    dt = parse_dt(group_date)
                    if dt is not None:
                        if best_dt is None or dt > best_dt or (dt == best_dt and valid_count >= best_valid):
                            best_dt = dt
                            best_valid = valid_count
                            best_idx = idx
                            flattened = items
                    else:
                        if best_dt is None and valid_count >= best_valid:
                            best_valid = valid_count
                            best_idx = idx
                            flattened = items

                return flattened if (best_idx is not None and flattened) else None

            if table_json:
                try:
                    tdata = json.loads(table_json)
                    pt_effective_dt = _extract_effective_date(tdata)
                    metal_list = tdata.get("metalList", tdata.get("prices", tdata.get("data", [])))

                    if not isinstance(metal_list, list):
                        metal_list = []

                    latest_group = pick_latest_valid_group(metal_list)
                    if latest_group:
                        if pt_effective_dt is None:
                            for it in latest_group:
                                if isinstance(it, dict):
                                    pt_effective_dt = _extract_effective_date(it)
                                    if pt_effective_dt:
                                        break

                        for item in latest_group:
                            if not isinstance(item, dict):
                                continue
                            name = _norm_market(item.get("marketName") or item.get("market") or item.get("marketNameEn") or "")
                            t = _norm_time(item.get("marketTime") or item.get("time") or item.get("fixTime") or "")
                            price = _to_float(item.get("price") or item.get("value") or item.get("currentPrice"))

                            # 时段匹配：支持 _norm_time 归一化后的 "08:30" 等，以及原始字符串包含 "09:00" 等
                            if name == "Hong Kong" and (t == "08:30" or "08:30" in str(item.get("marketTime") or item.get("time") or "") or "8:30" in str(item.get("marketTime") or item.get("time") or "")):
                                pt_time_prices["Hong Kong 08:30"] = price
                            elif name == "Hong Kong" and (t == "14:00" or "14:00" in str(item.get("marketTime") or item.get("time") or "") or "2:00" in str(item.get("marketTime") or item.get("time") or "")):
                                pt_time_prices["Hong Kong 14:00"] = price
                            elif name == "London" and (t == "09:00" or "09:00" in str(item.get("marketTime") or item.get("time") or "") or "9:00" in str(item.get("marketTime") or item.get("time") or "")):
                                pt_time_prices["London 09:00"] = price
                            elif name == "New York" and (t == "09:30" or "09:30" in str(item.get("marketTime") or item.get("time") or "") or "9:30" in str(item.get("marketTime") or item.get("time") or "")):
                                pt_time_prices["New York 09:30"] = price

                        # 若按名称未匹配全，则按顺序填：JM 常为 [HK 08:30, HK 14:00, London 09:00, NY 09:30]
                        if sum(1 for v in pt_time_prices.values() if v is not None) < 4 and len(latest_group) >= 4:
                            for i, col in enumerate(TIME_COLUMNS):
                                if pt_time_prices.get(col) is not None:
                                    continue
                                it = latest_group[i] if i < len(latest_group) and isinstance(latest_group[i], dict) else None
                                if it is not None:
                                    p = _to_float(it.get("price") or it.get("value") or it.get("currentPrice"))
                                    if p is not None:
                                        pt_time_prices[col] = p
                    pt_filled = sum(1 for v in pt_time_prices.values() if v is not None)
                    _log(f"Pt 分时价格: {pt_time_prices} (pt_effective_dt={pt_effective_dt})")
                    if pt_filled == 0 and table_json:
                        _log("[WARN] 铂金分时价格为 0 条，请检查页面结构；可设 DEBUG_SAVE_RAW_TABLE=True 后查看 PGM_debug_metalTablePrices.json")
                    elif pt_filled > 0 and pt_filled < 4:
                        _log(f"[INFO] 铂金仅解析到 {pt_filled}/4 个时段，若与网页不符可查看 PGM_debug_metalTablePrices.json 核对字段")
                except Exception as e:
                    _log(f"[WARN] 解析 metalTablePrices 失败: {e}")

            # 3) 调试保存 JSON
            if DEBUG_SAVE_RAW_TABLE:
                try:
                    dir_path = os.path.dirname(os.path.abspath(__file__))
                    if cur_json:
                        with open(os.path.join(dir_path, "PGM_debug_currentMetalPrices.json"), "w", encoding="utf-8") as f:
                            f.write(cur_json)
                        _log("已保存 PGM_debug_currentMetalPrices.json")
                    if table_json:
                        with open(os.path.join(dir_path, "PGM_debug_metalTablePrices.json"), "w", encoding="utf-8") as f:
                            f.write(table_json)
                        _log("已保存 PGM_debug_metalTablePrices.json")
                    meta = {
                        "session_override": session_override,
                        "cur_effective_dt": cur_effective_dt,
                        "pt_effective_dt": pt_effective_dt,
                        "pt_time_prices": pt_time_prices,
                    }
                    with open(os.path.join(dir_path, "PGM_debug_parsed_meta.json"), "w", encoding="utf-8") as f:
                        json.dump(meta, f, ensure_ascii=False, indent=2)
                    _log("已保存 PGM_debug_parsed_meta.json")
                except Exception as ex:
                    _log(f"[WARN] 保存调试 JSON 失败: {ex}")

            # 4) 组装 records
            records = []
            for metal in TARGET_METALS:
                # 行级 Date：与网站一致，优先用 JM 返回的有效日期；仅当接口未返回时用抓取日
                if USE_TODAY_AS_DATE:
                    row_date = today_hkt
                elif metal == "Platinum":
                    row_date = pt_effective_dt or cur_effective_dt or today_hkt
                else:
                    row_date = cur_effective_dt or today_hkt

                row_data = {
                    "Date": row_date,
                    "Metal": metal,
                    "Source": None,
                    "Hong Kong 08:30": None,
                    "Hong Kong 14:00": None,
                    "London 09:00": None,
                    "New York 09:30": None,
                }

                # currentMetalPrices 里找对应 metal
                def _metal_match(item: dict) -> bool:
                    n = item.get("metalName") or item.get("name") or item.get("metal")
                    return isinstance(n, str) and n == metal

                cur_items = [x for x in cur_list if isinstance(x, dict) and _metal_match(x)]
                base_price = None
                if cur_items:
                    p = cur_items[0].get("price") or cur_items[0].get("currentPrice") or cur_items[0].get("value")
                    base_price = _to_float(p)

                if metal == "Platinum":
                    # Pt：只用 tablePrices（Fix）
                    for col in TIME_COLUMNS:
                        if pt_time_prices.get(col) is not None:
                            row_data[col] = pt_time_prices[col]
                    if not any(row_data[c] is not None for c in TIME_COLUMNS):
                        # 没有任何分时价则不写入
                        continue
                    row_data["Source"] = "Fix"
                else:
                    # Pd/Rh：session_override 为 None 时（force 一键抓取）写入全部四列；否则只写对应时段列。
                    if base_price is None:
                        continue
                    if session_override is None:
                        for col in TIME_COLUMNS:
                            row_data[col] = base_price
                    else:
                        target_col = (session_override if session_override in TIME_COLUMNS else None) or _current_session_column()
                        if not target_col:
                            _log(f"[WARN] 无法确定写入列（session_override={session_override}），{metal} 本次不写入")
                            continue
                        row_data[target_col] = base_price
                    row_data["Source"] = "Current"

                records.append(row_data)

            if not records:
                return pd.DataFrame()

            df = pd.DataFrame(records)
            _log(f"解析结果:\n{df}")
            return df[["Date", "Metal", "Source"] + TIME_COLUMNS]
        finally:
            browser.close()


def _ensure_output_columns(df: pd.DataFrame) -> pd.DataFrame:
    """确保输出包含 Date/Metal/Source + TIME_COLUMNS。"""
    if "Source" not in df.columns:
        df["Source"] = None
    for c in TIME_COLUMNS:
        if c not in df.columns:
            df[c] = None
    return df[["Date", "Metal", "Source"] + TIME_COLUMNS]


def _add_trend_chart(ws, sheet_name: str) -> None:
    """
    在当前工作表中添加价格走势图。数据列：A=Date, B=Source, C~F=四时段价格。
    图表放在数据右侧（约 H2 起），仅当至少有 2 行数据时添加。
    """
    if ws.max_row < 2:
        return
    try:
        # 横轴：日期（A 列，第 2 行起）
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        # 纵轴：四列价格（C~F），含标题行
        data = Reference(ws, min_col=3, min_row=1, max_col=6, max_row=ws.max_row)
        chart = LineChart()
        chart.title = f"{sheet_name} 价格走势"
        chart.y_axis.title = "Price"
        chart.x_axis.title = "Date"
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 18
        chart.height = 10
        ws.add_chart(chart, "H2")
    except Exception as e:
        _log(f"[WARN] 为 Sheet [{sheet_name}] 添加走势图失败: {e}")


def _append_to_excel(new_df: pd.DataFrame) -> None:
    """
    写入 all_metals_data.xlsx：
    - 每金属一个 Sheet（Platinum/Palladium/Rhodium）
    - Sheet 列：Date, Source, TIME_COLUMNS
    - 合并键：Date + Source（同一天 Current 与 Fix 不覆盖）
    """
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FILE)

    # 至少要有一个时段价格非空才写
    price_cols = [c for c in TIME_COLUMNS if c in new_df.columns]
    if not price_cols:
        _log("[写入] 无时间列数据，跳过")
        return

    new_df = _ensure_output_columns(new_df)
    has_price = new_df[price_cols].notna().any(axis=1)
    new_df = new_df.loc[has_price].copy()
    if new_df.empty:
        _log("[写入] 无有效数据（所有价格为空），跳过")
        return

    _log(f"[写入] 本批 {len(new_df)} 行，写入 {file_path}")

    # 读取已有
    existing_sheets = {}
    if os.path.exists(file_path):
        try:
            all_sheets = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
            for name, df in all_sheets.items():
                if df is None or df.empty:
                    continue
                if "Date" not in df.columns:
                    continue
                df = df.copy()
                df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
                if "Source" not in df.columns:
                    df["Source"] = None
                for c in TIME_COLUMNS:
                    if c not in df.columns:
                        df[c] = None
                # 每个金属每天只保留 1 行：同日期多行合并为一行，四列取最后非空
                out_cols = ["Date", "Source"] + TIME_COLUMNS
                dff = df[out_cols].copy()

                def _last_valid(s):
                    x = s.dropna()
                    return x.iloc[-1] if len(x) else None

                agg_dict = {c: _last_valid for c in TIME_COLUMNS}
                agg_dict["Source"] = "last"
                dff = dff.groupby("Date", as_index=False).agg(agg_dict)
                existing_sheets[str(name).strip()] = dff[[c for c in out_cols if c in dff.columns]]
        except Exception as e:
            _log(f"[WARN] 读取已有 Excel 失败，将新建: {e}")

    def _norm_date(d) -> str:
        if d is None or (isinstance(d, float) and pd.isna(d)):
            return ""
        if hasattr(d, "strftime"):
            return d.strftime("%Y-%m-%d")
        s = str(d).strip()
        if not s or s == "NaT":
            return ""
        try:
            out = pd.to_datetime(s, errors="coerce")
            if pd.isna(out):
                return s[:10] if len(s) >= 10 else s
            return out.strftime("%Y-%m-%d")
        except Exception:
            return s[:10] if len(s) >= 10 else s

    def merge_sheet(existing: pd.DataFrame, new_part: pd.DataFrame) -> pd.DataFrame:
        """每个金属每天只保留 1 行（Date 唯一键）；同一天新数据只覆盖对应时段列。"""
        out_cols = ["Date", "Source"] + TIME_COLUMNS
        if existing is None or existing.empty:
            tmp = new_part.copy()
            tmp["Date"] = tmp["Date"].apply(_norm_date)
            if "Source" not in tmp.columns:
                tmp["Source"] = None
            return tmp[out_cols].drop_duplicates(subset=["Date"], keep="last").sort_values("Date", ascending=False)

        ex = existing.copy()
        ex["Date"] = ex["Date"].apply(_norm_date)
        if "Source" not in ex.columns:
            ex["Source"] = None
        ex = ex[out_cols].drop_duplicates(subset=["Date"], keep="last")

        npart = new_part.copy()
        npart["Date"] = npart["Date"].apply(_norm_date)

        for _, row in npart.iterrows():
            d = row.get("Date", "")
            if not d:
                continue
            mask = ex["Date"].astype(str) == str(d)
            if mask.any():
                idx = ex.index[mask][0]
                for c in TIME_COLUMNS:
                    v = row.get(c)
                    if pd.notna(v):
                        ex.at[idx, c] = v
                if "Source" in row.index and pd.notna(row.get("Source")):
                    ex.at[idx, "Source"] = row.get("Source")
            else:
                new_row = {col: row.get(col) for col in out_cols}
                ex = pd.concat([ex, pd.DataFrame([new_row])], ignore_index=True)

        return ex.drop_duplicates(subset=["Date"], keep="last").sort_values("Date", ascending=False)

    # 合并并写出
    sheets_to_write = {}
    for metal in TARGET_METALS:
        existing = existing_sheets.get(metal, pd.DataFrame(columns=["Date", "Source"] + TIME_COLUMNS))
        new_part = new_df[new_df["Metal"] == metal][["Date", "Source"] + TIME_COLUMNS]
        sheets_to_write[metal] = merge_sheet(existing, new_part) if not new_part.empty else existing

    dir_path = os.path.dirname(file_path)
    base_name = os.path.splitext(OUTPUT_FILE)[0]
    tmp_path = os.path.join(dir_path, base_name + "_tmp.xlsx")

    try:
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            for metal in TARGET_METALS:
                sheets_to_write[metal].to_excel(writer, sheet_name=metal, index=False)
            # 为每个 Sheet 添加走势图（数据列 A=Date, B=Source, C~F=四时段）
            for metal in TARGET_METALS:
                ws = writer.book[metal]
                _add_trend_chart(ws, metal)

        if os.path.exists(file_path):
            os.remove(file_path)
        os.rename(tmp_path, file_path)
        _log("[写入] 完成写入并覆盖旧文件（如 Excel 打开会失败，请先关闭）")
    except PermissionError:
        _log(f"[ERROR] 写入失败：文件可能正被打开：{file_path}")
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        raise
    except Exception as e:
        _log(f"[ERROR] 写入失败：{e}")
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        raise


def _do_scrape_once(session_override: Optional[str] = None) -> Tuple[bool, pd.DataFrame]:
    """执行一次抓取并写入 Excel。"""
    global _latest_missing_items
    session_col = session_override or _current_session_column()

    try:
        cleaned = _scrape_rows_with_playwright(session_col)
        if cleaned is None or cleaned.empty:
            with _state_lock:
                _latest_missing_items = []
            return False, pd.DataFrame()

        # 缺失项记录：对每金属统计哪些列全为空
        missing = []
        for metal in TARGET_METALS:
            mdf = cleaned[cleaned["Metal"] == metal]
            if mdf.empty:
                for col in TIME_COLUMNS:
                    missing.append(f"{metal} - {col}")
                continue
            for col in TIME_COLUMNS:
                if mdf[col].isna().all():
                    missing.append(f"{metal} - {col}")

        with _state_lock:
            _latest_missing_items = missing

        with _excel_lock:
            _append_to_excel(cleaned)

        return True, cleaned
    except Exception as e:
        _log(f"[ERROR] 抓取失败: {e}")
        return False, pd.DataFrame()


def _scraper_worker() -> None:
    """后台线程：按调度时间执行抓取；若数据为空则每 5 分钟重试，最多等到该时段 1 小时后放弃。"""
    global _latest_scraped_record, _next_run_time
    while not _shutdown_event.is_set():
        next_t = _get_next_run_time_hkt()
        with _state_lock:
            _next_run_time = next_t

        now = datetime.now(ZoneInfo("Asia/Hong_Kong"))
        while now < next_t and not _shutdown_event.is_set():
            time.sleep(1)
            now = datetime.now(ZoneInfo("Asia/Hong_Kong"))

        if _shutdown_event.is_set():
            break

        session_col = _session_name_for_run_time(next_t) or _current_session_column()
        success, cleaned = _do_scrape_once(session_col)

        if success and not cleaned.empty:
            with _state_lock:
                _latest_scraped_record = cleaned.iloc[-1].to_dict()
            continue

        # 重试
        session_start = next_t
        deadline = session_start + timedelta(hours=WAIT_DEADLINE_HOURS)
        while not _shutdown_event.is_set():
            now = datetime.now(ZoneInfo("Asia/Hong_Kong"))
            if now >= deadline:
                break

            _log("等待官网更新中...")
            time.sleep(10)

            # 每 WAIT_RETRY_MINUTES 再抓一次
            if (deadline - now).total_seconds() <= 0:
                break

            # 控制重试节奏
            # 简化：每 WAIT_RETRY_MINUTES 触发一次
            if now.minute % WAIT_RETRY_MINUTES != 0:
                continue

            success, cleaned = _do_scrape_once(session_col)
            if success and not cleaned.empty:
                with _state_lock:
                    _latest_scraped_record = cleaned.iloc[-1].to_dict()
                break


def _cmd_help() -> None:
    print(
        """
指令说明
--------
status  查看运行状态（香港时间、最近抓取记录、缺失项、下次抓取倒计时）
view    用默认程序打开 all_metals_data.xlsx
force   一键抓取全部地区/时段，无需选择
exit    安全退出程序
help    显示本说明
""".strip()
    )


def _cmd_status() -> None:
    with _state_lock:
        latest = _latest_scraped_record
        next_t = _next_run_time
        missing = list(_latest_missing_items) if _latest_missing_items else []

    now = datetime.now(ZoneInfo("Asia/Hong_Kong"))
    _log(f"当前系统时间（香港）: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    _log(f"最近抓取记录: {latest if latest else '（暂无）'}")
    _log(f"本次抓取缺失项: {', '.join(missing) if missing else '无'}")

    if next_t:
        delta = next_t - now
        if delta.total_seconds() > 0:
            _log(f"下次抓取倒计时: {str(delta).split('.')[0]}")
        else:
            _log("下次抓取倒计时: 即将执行")
    else:
        _log("下次抓取倒计时: （未设定）")


def _cmd_view() -> None:
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FILE)
    if not os.path.exists(file_path):
        _log(f"Excel 文件不存在: {file_path}")
        return
    file_path = os.path.abspath(file_path)
    try:
        if sys.platform == "win32":
            os.startfile(file_path)  # type: ignore
        elif sys.platform == "darwin":
            subprocess.run(["open", file_path], check=False)
        else:
            subprocess.run(["xdg-open", file_path], check=False)
        _log(f"已打开: {file_path}")
    except Exception as e:
        _log(f"打开 Excel 失败: {e}")


def _cmd_force() -> None:
    global _latest_scraped_record
    _log("立即执行抓取（全部地区/时段，无需选择）...")
    success, cleaned = _do_scrape_once(None)
    if success:
        _log("抓取完成并已写入")
        with _state_lock:
            _latest_scraped_record = cleaned.iloc[-1].to_dict() if not cleaned.empty else None
    else:
        _log("本次抓取无有效数据")


def run() -> None:
    """
    单次抓取（--once）：
    统一走 _do_scrape_once（Playwright JSON 解析链路），避免旧 run() 走 HTML table 导致数据/结构不一致。
    """
    _log("---------- 开始执行 JM PGM 抓取（单次模式） ----------")
    _log("当前时间（香港）: %s" % datetime.now(ZoneInfo("Asia/Hong_Kong")).strftime("%Y-%m-%d %H:%M:%S"))

    session_col = _current_session_column()
    if not session_col:
        _log("[WARN] 当前不在任何时段窗口内（HK08:30/HK14:00/LDN09:00/NY09:30）。Pd/Rh 可能不会写入任何列。")

    success, cleaned = _do_scrape_once(session_col)
    if not success or cleaned.empty:
        _log("[WARN] 本次抓取未获得有效数据（可能 JM 未更新/网络/结构变更）")
        return

    _log("---------- 单次抓取完成，已写入 Excel ----------")


def run_interactive() -> None:
    """多线程交互模式：后台定时抓取 + 前台 input() 指令。"""
    global _next_run_time
    _log("---------- JM PGM 抓取（多线程交互模式）----------")
    _log("输入 help 查看指令说明；输入 exit 退出")

    with _state_lock:
        _next_run_time = _get_next_run_time_hkt()

    scraper = threading.Thread(target=_scraper_worker, daemon=False)
    scraper.start()

    try:
        while True:
            try:
                cmd = input("输入指令 (status/view/force/exit/help): ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                break

            if cmd == "status":
                _cmd_status()
            elif cmd == "view":
                _cmd_view()
            elif cmd == "force":
                _cmd_force()
            elif cmd == "help":
                _cmd_help()
            elif cmd == "exit":
                break
            elif cmd:
                _log("未知指令，输入 help 查看用法")
    finally:
        _log("正在安全退出...")
        _shutdown_event.set()
        scraper.join(timeout=20)
        _log("已退出")
    sys.exit(0)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1].strip() == "--once":
        run()
    else:
        run_interactive()